"""
ETL: Parse all health data sources into normalized CSVs.
Run from repo root: python scripts/etl.py
"""
import json
import glob
import os
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = DATA / "processed"
OUT.mkdir(exist_ok=True)


def parse_garmin_daily():
    """Parse Garmin daily JSON files into a flat DataFrame."""
    records = []
    for fp in sorted(glob.glob(str(DATA / "garmin" / "*.json"))):
        with open(fp, "r", encoding="utf-8") as f:
            d = json.load(f)
        row = {"date": d["date"]}
        # Sleep
        sl = d.get("sleep", {})
        for k in ["score", "total_hours", "deep_hours", "light_hours", "rem_hours", "awake_min"]:
            row[f"sleep_{k}"] = sl.get(k)
        # HRV
        hrv = d.get("hrv", {})
        row["hrv_last_night_avg"] = hrv.get("last_night_avg")
        row["hrv_weekly_avg"] = hrv.get("weekly_avg")
        # Stress
        st = d.get("stress", {})
        row["stress_overall"] = st.get("overall")
        # Body Battery
        bb = d.get("body_battery", {})
        row["body_battery_charged"] = bb.get("charged")
        row["body_battery_drained"] = bb.get("drained")
        # Activity
        act = d.get("activity", {})
        for k in ["steps", "distance_km", "active_calories", "resting_hr", "max_hr",
                   "moderate_intensity_min", "vigorous_intensity_min", "floors_climbed"]:
            row[f"activity_{k}"] = act.get(k)
        records.append(row)
    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    df.to_csv(OUT / "garmin_daily.csv", index=False)
    print(f"  garmin_daily: {len(df)} days ({df['date'].min().date()} to {df['date'].max().date()})")
    return df


def parse_garmin_activities():
    """Parse Garmin activities CSV."""
    fp = DATA / "activities" / "activities.csv"
    df = pd.read_csv(fp)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    df.to_csv(OUT / "activities.csv", index=False)
    print(f"  activities: {len(df)} workouts ({df['date'].min().date()} to {df['date'].max().date()})")
    return df


def parse_dexa():
    """Parse DEXA Excel into normalized rows."""
    wb = openpyxl.load_workbook(DATA / "dexa" / "DEXA_RESULTS.xlsx")
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        scan_date = rows[0][1]  # datetime in B1
        if isinstance(scan_date, datetime):
            scan_date = scan_date.strftime("%Y-%m-%d")
        height = rows[2][2]  # C3
        weight = rows[3][2]  # C4

        rec = {"date": scan_date, "height_cm": height, "weight_kg": weight}

        # Body composition summary (columns I-O, rows 6+)
        # TOTAL row: row index 12 (0-based) = row 13
        for r in rows[5:]:
            region = r[8]  # col I
            if region == "TOTAL":
                rec["total_fat_g"] = r[9]
                rec["total_lean_bmc_g"] = r[10]
                rec["total_mass_g"] = r[11]
                rec["total_pct_fat"] = r[12]
            elif region == "TRUNK":
                rec["trunk_fat_g"] = r[9]
                rec["trunk_lean_bmc_g"] = r[10]
                rec["trunk_pct_fat"] = r[12]
            elif region == "ANDROID (A)":
                rec["android_fat_g"] = r[9]
                rec["android_pct_fat"] = r[12]
            elif region == "GYNOID (G)":
                rec["gynoid_fat_g"] = r[9]
                rec["gynoid_pct_fat"] = r[12]

        # Adipose indices (columns Q-T)
        for r in rows[5:]:
            measure = r[16]  # col Q
            result = r[17]   # col R
            if measure == "TOTAL BODY % FAT":
                rec["body_pct_fat_adipose"] = result
            elif measure == "FAT MASS/HEIGHT^2 (kg/m2)":
                rec["fmi"] = result
            elif measure == "ANDROID/GYNOID RATIO":
                rec["ag_ratio"] = result
            elif measure == "EST. VAT MASS (g)":
                rec["vat_mass_g"] = result
            elif measure == "EST. VAT AREA (cm2)":
                rec["vat_area_cm2"] = result
            elif measure and "LEAN/HEIGHT" in str(measure) and "APPEN" not in str(measure):
                rec["lmi"] = result
            elif measure and "APPEN" in str(measure):
                rec["almi"] = result

        # BMD from bone summary (columns B-G)
        for r in rows[5:]:
            region = r[1]
            if region == "TOTAL":
                rec["bmd_total"] = r[4] if r[4] != "=SUM(C6:C16)-C15" else None
                rec["t_score"] = r[5]
                rec["z_score"] = r[6]
            elif region == "L SPINE":
                rec["bmd_l_spine"] = r[4]
            elif region == "PELVIS":
                rec["bmd_pelvis"] = r[4]

        records.append(rec)

    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    df.to_csv(OUT / "dexa.csv", index=False)
    print(f"  dexa: {len(df)} scans ({df['date'].min().date()} to {df['date'].max().date()})")
    return df


def parse_blood_tests():
    """Parse blood test Excel into long-format DataFrame."""
    wb = openpyxl.load_workbook(DATA / "blood_tests" / "BLOODTESTRESULTS.xlsx")
    ws = wb[wb.sheetnames[0]]

    # Row 2 has dates in columns D, E, F, G (indices 3-6)
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    dates = []
    for i in range(3, 7):
        d = header_row[i]
        if isinstance(d, datetime):
            dates.append(d.strftime("%Y-%m-%d"))
        else:
            dates.append(str(d))

    records = []
    current_category = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] is not None:
            current_category = str(row[1])
        marker = row[2]
        if marker is None:
            continue
        marker = str(marker).strip()
        # Clean Korean characters from marker names
        clean_marker = marker.split("(")[0].strip() if "(" in marker else marker
        for i, date in enumerate(dates):
            val = row[3 + i]
            if val is not None:
                records.append({
                    "date": date,
                    "category": current_category,
                    "marker": marker,
                    "marker_clean": clean_marker,
                    "value": val,
                    "is_numeric": isinstance(val, (int, float)),
                })

    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values(["marker", "date"]).reset_index(drop=True)
    df.to_csv(OUT / "blood_tests.csv", index=False)
    print(f"  blood_tests: {len(df)} measurements across {df['date'].nunique()} dates")
    return df


def parse_grip_strength():
    """Parse grip strength CSV."""
    fp = DATA / "grip_strength" / "grip_strength.csv"
    df = pd.read_csv(fp)
    df["date"] = pd.to_datetime(df["date"])
    df.to_csv(OUT / "grip_strength.csv", index=False)
    print(f"  grip_strength: {len(df)} measurements")
    return df


def parse_nback():
    """Parse N-back cognitive test sessions."""
    fp = DATA / "nback" / "sessions.csv"
    df = pd.read_csv(fp)
    df["datetime"] = pd.to_datetime(df["datetime"])
    df["date"] = df["datetime"].dt.date
    df.to_csv(OUT / "nback.csv", index=False)
    print(f"  nback: {len(df)} sessions")
    return df


def main():
    print("=== Health Tracker ETL ===")
    parse_garmin_daily()
    parse_garmin_activities()
    parse_dexa()
    parse_blood_tests()
    parse_grip_strength()
    parse_nback()
    print("=== Done. Processed CSVs in data/processed/ ===")


if __name__ == "__main__":
    main()
